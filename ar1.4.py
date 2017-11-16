#!/usr/bin/python
#C:\pythontests\autorep\ar1.4.py
from shutil import copyfile
from openpyxl import load_workbook
from stat import S_ISREG, ST_CTIME, ST_MODE
import os, openpyxl, csv, re, sys, time, datetime, glob

rootDir = "S:/O&M/AA_reporting/2017_MonthlyReports"
refDirlist = glob.glob('S:/O&M/AA_reporting/*.xlsx') # * means all if need specific format then *.csv

refwb = max(refDirlist, key=os.path.getctime)
refwb = load_workbook(refwb, data_only=True)
limit = datetime.datetime.now().strftime('%m-%d-%Y').split("-")
limit = limit[2]+"_"+limit[0]

for subdir in os.listdir(rootDir):
    w=3
    x=3
    y=2
    z=3
    filelist = glob.glob(rootDir+'/'+subdir+'/'+'*.xlsx')
    oldwb = min(filelist, key=os.path.getctime)
    newwb = load_workbook(oldwb)
    ref_sheet = refwb.get_sheet_by_name('mon_prod')
    new_sheet = newwb.get_sheet_by_name('Data')
    id_old = ref_sheet['A%a' %x]
    id_new = new_sheet['A%a' %y]
    name_old = ref_sheet['B%a' %x]
    name_new = new_sheet['B%a' %y]
    refval = ref_sheet.cell(row=4,column=z)

    while id_new.value:
        x=3
        while id_old.value != id_new.value and name_old.value != name_new.value:
            #print ('or here?')
            z=4
            w=3
            x=x+1
            id_old = ref_sheet['A%a' %x]
            id_new = new_sheet['A%a' %y]
            name_old = ref_sheet['B%a' %x]
            name_new = new_sheet['B%a' %y]
        while ref_sheet.cell(row=1,column=z).value != '2016_01':
            z=z+1
            #refval = ref_sheet.cell(row=x,column=z)
            print (refval.value)
        while ref_sheet.cell(row=1,column=z).value != limit and z != 1000:
            new_sheet.cell(row=y,column=w).value = ref_sheet.cell(row=x,column=z).value
            new_sheet.cell(row=(1),column=w).value = ref_sheet.cell(row=1,column=z).value
            w=w+1
            z=z+1
            print ((ref_sheet.cell(row=1,column=z)).value)
            print ((ref_sheet.cell(row=1,column=z)))
        y=y+1
        id_new = new_sheet['A%a' %y]
        name_new = new_sheet['B%a' %y]


    newpath = oldwb.split('.')
    newpath = newpath[0]+datetime.datetime.now().strftime('_%m-%d-%y')+".xlsx"

    print (newpath)
    newwb.save(newpath)
